"""Independent reference implementation of the council's calculation engine.

This exists to disagree. The browser engine computes every figure twice already
-- once in SQL, once in a JavaScript reducer -- but both run in the same
process, written by the same hand, against the same in-memory rows. A third
implementation in a different language, structured differently, is what turns
"the two agree" into "the number is right".

Nothing here is specific to any dataset. The grain, the flow/stock split, and
the incomplete periods are all discovered from the workbook's own structure, in
the same way `contract.js` discovers them, but by a different route.

**No expected figure appears in this file.** Every acceptance value lives in
`tests/fixtures/*.json` and nowhere else.

    python reference.py --workbook path/to/book.xlsx
    python reference.py --workbook path/to/book.xlsx --json
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Any, Callable, Iterable, Sequence

try:
    import openpyxl
except ImportError:  # pragma: no cover - surfaced by the CLI and by pytest
    openpyxl = None


DAY = dt.timedelta(days=1)
MATCHED_OFFSET_DAYS = 364  # 52 whole weeks: keeps weekday alignment intact


# --------------------------------------------------------------------------
# loading
# --------------------------------------------------------------------------

@dataclass
class Sheet:
    name: str
    header: list[str]
    rows: list[list[Any]]

    def col(self, name: str) -> int:
        return self.header.index(name)

    def values(self, name: str) -> list[Any]:
        i = self.col(name)
        return [r[i] for r in self.rows]


def read_workbook(path: str) -> dict[str, Sheet]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required: pip install -r requirements.txt")
    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[str, Sheet] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(c not in (None, "") for c in r)]
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        width = len(header)
        body = [list(r[:width]) + [None] * max(0, width - len(r)) for r in rows[1:]]
        out[name] = Sheet(name, header, body)
    return out


# --------------------------------------------------------------------------
# typing helpers
# --------------------------------------------------------------------------

def as_num(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def as_date(v: Any) -> str | None:
    if isinstance(v, (dt.date, dt.datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        s = v.strip()[:10]
        try:
            dt.date.fromisoformat(s)
            return s
        except ValueError:
            return None
    return None


def column_kind(values: Sequence[Any]) -> str:
    filled = [v for v in values if v not in (None, "")]
    if not filled:
        return "empty"
    dates = sum(1 for v in filled if as_date(v) is not None)
    if dates / len(filled) > 0.95:
        return "date"
    nums = sum(1 for v in filled if as_date(v) is None and as_num(v) is not None)
    if nums / len(filled) > 0.95:
        return "number"
    return "string"


def shift_days(iso: str, days: int) -> str:
    return (dt.date.fromisoformat(iso) + dt.timedelta(days=days)).isoformat()


# --------------------------------------------------------------------------
# the data contract, rediscovered
# --------------------------------------------------------------------------

@dataclass
class Contract:
    grain: list[str]
    grain_unique: bool
    demoted_labels: list[str]
    period_col: str | None
    coverage_dim: str | None
    incomplete_periods: list[str]
    roles: dict[str, str]                  # column -> flow | stock | attribute
    split_groups: int
    duplicate_groups: int
    raw_rows: int
    collapsed_rows: int
    extras: dict[str, Any] = field(default_factory=dict)


def _is_label_of(sheet: Sheet, coarse: str, fine: str) -> bool:
    """Is `coarse` a label riding along with `fine` rather than its own axis?

    Measured as the mean number of distinct `coarse` values per `fine` value.
    A calendar label sits at ~1.0 (a week belongs to one month, or to two when
    it straddles a boundary). A genuine key part sits at the cardinality of the
    dimension, because every one of its members recurs in every period.
    """
    ci, fi = sheet.col(coarse), sheet.col(fine)
    if len({str(r[ci]) for r in sheet.rows}) >= len({str(r[fi]) for r in sheet.rows}):
        return False
    buckets: dict[str, set[str]] = defaultdict(set)
    for r in sheet.rows:
        buckets[str(r[fi])].add(str(r[ci]))
    if not buckets:
        return False
    return statistics.fmean(len(v) for v in buckets.values()) < 1.5


def _key_of(row: Sequence[Any], idx: Sequence[int]) -> tuple:
    return tuple("" if row[i] is None else str(row[i]) for i in idx)


def profile(sheet: Sheet) -> Contract:
    kinds = {c: column_kind(sheet.values(c)) for c in sheet.header}
    n = len(sheet.rows)

    candidates = [
        c for c in sheet.header
        if kinds[c] in ("date", "string")
        and 1 < len({str(v) for v in sheet.values(c)}) < n
    ]
    candidates.sort(key=lambda c: -len({str(v) for v in sheet.values(c)}))

    demoted = [
        c for c in candidates
        if any(f != c and _is_label_of(sheet, c, f) for f in candidates)
    ]
    key_pool = [c for c in candidates if c not in demoted] or candidates

    grain: list[str] = []
    unique = False
    for c in key_pool:
        grain.append(c)
        idx = [sheet.col(g) for g in grain]
        if len({_key_of(r, idx) for r in sheet.rows}) == n:
            unique = True
            break

    idx = [sheet.col(g) for g in grain]
    groups: dict[tuple, list[int]] = defaultdict(list)
    for i, r in enumerate(sheet.rows):
        groups[_key_of(r, idx)].append(i)

    non_key = [c for c in sheet.header if c not in grain]
    split = dup = 0
    multi: list[list[int]] = []
    for members in groups.values():
        if len(members) < 2:
            continue
        differs = any(
            len({str(sheet.rows[i][sheet.col(c)]) for i in members}) > 1
            for c in non_key
        )
        if differs:
            split += 1
            multi.append(members)
        else:
            dup += 1

    # flow vs stock, from how segments of one key behave
    roles: dict[str, str] = {}
    for c in sheet.header:
        if c in grain:
            continue
        if kinds[c] != "number":
            roles[c] = "attribute"
            continue
        ratios: list[float] = []
        for members in multi:
            vals = [as_num(sheet.rows[i][sheet.col(c)]) for i in members]
            vals = [v for v in vals if v is not None and v != 0]
            if len(vals) < 2:
                continue
            ratios.append(max(vals) / min(vals))
        if ratios and 0.9 <= statistics.median(ratios) <= 1.15:
            roles[c] = "stock"     # one level recorded twice -> never sum it
        else:
            roles[c] = "flow"
    period_col = next((c for c in grain if kinds[c] == "date"), None)

    # the steadiest reporting dimension decides what "complete" means
    coverage_dim = None
    incomplete: list[str] = []
    if period_col:
        pi = sheet.col(period_col)
        best_cv = None
        for c in grain:
            if c == period_col or kinds[c] != "string":
                continue
            per: dict[str, set[str]] = defaultdict(set)
            for r in sheet.rows:
                per[str(as_date(r[pi]) or r[pi])].add(str(r[sheet.col(c)]))
            counts = [len(v) for v in per.values()]
            if len(counts) < 2:
                continue
            mean = statistics.fmean(counts)
            cv = (statistics.pstdev(counts) / mean) if mean else float("inf")
            if best_cv is None or cv < best_cv:
                best_cv, coverage_dim = cv, c
        if coverage_dim:
            per = defaultdict(set)
            for r in sheet.rows:
                per[str(as_date(r[pi]) or r[pi])].add(str(r[sheet.col(coverage_dim)]))
            modal = Counter(len(v) for v in per.values()).most_common(1)[0][0]
            incomplete = sorted(p for p, v in per.items() if len(v) < modal)

    # How many distinct periods are touched by at least one split group —
    # the blast radius of the boundary problem, not just its row count.
    split_periods: set[str] = set()
    if period_col:
        pi = sheet.col(period_col)
        for members in multi:
            split_periods.add(as_date(sheet.rows[members[0]][pi]) or str(sheet.rows[members[0]][pi]))

    return Contract(
        grain=grain, grain_unique=unique, demoted_labels=demoted,
        period_col=period_col, coverage_dim=coverage_dim,
        incomplete_periods=incomplete, roles=roles,
        split_groups=split, duplicate_groups=dup,
        raw_rows=n, collapsed_rows=len(groups),
        extras={"splitPeriods": sorted(split_periods)},
    )


# --------------------------------------------------------------------------
# collapse
# --------------------------------------------------------------------------

@dataclass
class Fact:
    period: str
    dims: dict[str, str]
    measures: dict[str, float]
    segments: int


def collapse(sheet: Sheet, contract: Contract) -> list[Fact]:
    idx = [sheet.col(g) for g in contract.grain]
    groups: dict[tuple, list[int]] = defaultdict(list)
    for i, r in enumerate(sheet.rows):
        groups[_key_of(r, idx)].append(i)

    out: list[Fact] = []
    for key in sorted(groups):
        members = groups[key]
        measures: dict[str, float] = {}
        for col, role in contract.roles.items():
            if role == "attribute":
                continue
            vals = [as_num(sheet.rows[i][sheet.col(col)]) for i in members]
            vals = [v for v in vals if v is not None]
            if not vals:
                continue
            # A flow accumulates across its segments; a stock is one level
            # recorded once per segment, so the fold must not add them up.
            measures[col] = sum(vals) if role == "flow" else max(vals)
        dims = dict(zip(contract.grain, key))
        period = ""
        if contract.period_col:
            period = as_date(dims[contract.period_col]) or dims[contract.period_col]
            dims = {k: v for k, v in dims.items() if k != contract.period_col}
        out.append(Fact(period=period, dims=dims, measures=measures, segments=len(members)))
    return out


# --------------------------------------------------------------------------
# measures
# --------------------------------------------------------------------------

class Engine:
    def __init__(self, facts: list[Fact], contract: Contract,
                 weights: dict[str, dict[str, float]] | None = None,
                 weight_dim: str | None = None):
        self.contract = contract
        self.weight_dim = weight_dim
        self.weights = weights or {}
        bad = set(contract.incomplete_periods)
        self.facts = [f for f in facts if f.period not in bad]
        self.excluded = [f for f in facts if f.period in bad]
        self.periods = sorted({f.period for f in self.facts})

    # -- selection ---------------------------------------------------------
    def window(self, start: str, end: str,
               where: Callable[[Fact], bool] | None = None) -> Iterable[Fact]:
        for f in self.facts:
            if start <= f.period <= end and (where is None or where(f)):
                yield f

    def total(self, measure: str, start: str, end: str,
              where: Callable[[Fact], bool] | None = None,
              weight: str | None = None) -> float:
        acc = 0.0
        for f in self.window(start, end, where):
            v = f.measures.get(measure)
            if v is None:
                continue
            if weight:
                key = f.dims.get(self.weight_dim, "")
                w = self.weights.get(weight, {}).get(key)
                if w is None:
                    continue
                v *= w
            acc += v
        return acc

    def matched(self, measure: str, start: str, end: str,
                where: Callable[[Fact], bool] | None = None,
                weight: str | None = None) -> tuple[float, float, float | None]:
        cur = self.total(measure, start, end, where, weight)
        pa, pb = shift_days(start, -MATCHED_OFFSET_DAYS), shift_days(end, -MATCHED_OFFSET_DAYS)
        pri = self.total(measure, pa, pb, where, weight)
        return cur, pri, (cur / pri - 1) if pri else None

    def trailing(self, n: int) -> tuple[str, str]:
        return self.periods[-n], self.periods[-1]

    def at(self, measure: str, period: str,
           where: Callable[[Fact], bool] | None = None) -> float:
        return self.total(measure, period, period, where)

    # -- event time --------------------------------------------------------
    def event_matched(self, measure: str, anchors: dict[str, str], n: int,
                      entity_dim: str, before: bool = False,
                      weight: str | None = None) -> float | None:
        """Align each entity to its OWN anchor period, then compare.

        Calendar alignment answers "what happened this quarter". Event-time
        alignment answers "what happened in the first n periods after each
        entity changed" -- a different question, and the right one when a
        rollout is staggered.
        """
        cur = pri = 0.0
        for entity, anchor in anchors.items():
            if anchor not in self.periods:
                continue
            i = self.periods.index(anchor)
            span = self.periods[max(0, i - n):i] if before else self.periods[i:i + n]
            if not span:
                continue
            prior = {shift_days(p, -MATCHED_OFFSET_DAYS) for p in span}
            span_set = set(span)
            for f in self.facts:
                if f.dims.get(entity_dim) != entity:
                    continue
                v = f.measures.get(measure)
                if v is None:
                    continue
                if weight:
                    w = self.weights.get(weight, {}).get(f.dims.get(self.weight_dim, ""))
                    if w is None:
                        continue
                    v *= w
                if f.period in span_set:
                    cur += v
                elif f.period in prior:
                    pri += v
        return (cur / pri - 1) if pri else None

    def first_period_where(self, entity_dim: str, entity: str,
                           test: Callable[[Fact], bool]) -> str | None:
        hits = [f.period for f in self.facts
                if f.dims.get(entity_dim) == entity and test(f)]
        return min(hits) if hits else None


def build(path: str) -> tuple[Engine, Sheet, Contract, dict[str, Sheet]]:
    sheets = read_workbook(path)
    fact_sheet = max(sheets.values(), key=lambda s: len(s.rows))
    contract = profile(fact_sheet)
    facts = collapse(fact_sheet, contract)

    # An attribute table is a small sheet whose key column covers a fact column.
    weights: dict[str, dict[str, float]] = {}
    weight_dim = None
    for s in sheets.values():
        if s is fact_sheet or not (1 < len(s.rows) <= 200):
            continue
        for rc in s.header:
            ref_vals = {str(v) for v in s.values(rc) if v not in (None, "")}
            if len(ref_vals) < 2:
                continue
            for fc in contract.grain:
                fact_vals = {str(v) for v in fact_sheet.values(fc) if v not in (None, "")}
                if not fact_vals or len(fact_vals) > len(ref_vals) * 4:
                    continue
                if len(fact_vals & ref_vals) / len(fact_vals) > 0.95:
                    weight_dim = fc
                    ki = s.col(rc)
                    for col in s.header:
                        if col == rc:
                            continue
                        vals = [as_num(r[s.col(col)]) for r in s.rows]
                        if any(v is None for v in vals):
                            continue
                        weights[col] = {str(r[ki]): as_num(r[s.col(col)]) for r in s.rows}
                    break
            if weight_dim:
                break
        if weight_dim:
            break

    return Engine(facts, contract, weights, weight_dim), fact_sheet, contract, sheets


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def workbook_path(explicit: str | None = None) -> str | None:
    p = explicit or os.environ.get("COUNCIL_WORKBOOK")
    return p if p and os.path.exists(p) else None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", help="path to the source workbook")
    ap.add_argument("--json", action="store_true", help="emit the contract as JSON")
    args = ap.parse_args()

    path = workbook_path(args.workbook)
    if not path:
        print("No workbook. Pass --workbook or set COUNCIL_WORKBOOK.")
        return 2

    engine, sheet, contract, sheets = build(path)
    payload = {
        "workbook": os.path.basename(path),
        "sheets": list(sheets),
        "factSheet": sheet.name,
        "grain": contract.grain,
        "grainUnique": contract.grain_unique,
        "demotedLabels": contract.demoted_labels,
        "rawRows": contract.raw_rows,
        "collapsedRows": contract.collapsed_rows,
        "splitGroups": contract.split_groups,
        "duplicateGroups": contract.duplicate_groups,
        "roles": contract.roles,
        "periodColumn": contract.period_col,
        "coverageDimension": contract.coverage_dim,
        "incompletePeriods": contract.incomplete_periods,
        "completePeriods": len(engine.periods),
        "weightColumns": sorted(engine.weights),
        "weightDimension": engine.weight_dim,
    }
    if args.json:
        print(json.dumps(payload, indent=2))
    else:
        for k, v in payload.items():
            print(f"{k:>20}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
