"""Recomputes the Meridian traps from the committed workbook and asserts each
one is really there. Every check reads retention-demo.xlsx only - never the
generator's internal state - so it proves what an auditor could prove.

  (a) the deck's naive churn delta is ~2.1pt of "improvement"
  (b) a fixed-weight (pre-period lineage mix) delta is materially smaller
  (c) EMEA-UK / UK&I are one entity under two keys: complementary, continuous
  (d) Subscribers (EOM) is a level (non-additive across billing files) while
      the churn/addition flows are additive - proven via the ledger identity
  (e) the final month is structurally short of contributors vs every prior month

Also: the brief carries exactly one instruction-like paragraph for the
sentinel, and generate.py is deterministic (two fresh runs, identical bytes).
"""
import io, os, re, subprocess, sys, tempfile, zipfile
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "retention-demo.xlsx")
BRIEF = os.path.join(HERE, "brief.md")

MIGRATION = "2025-03-31"      # migration month (excluded by the deck)
FINAL = "2025-07-31"          # partial month
RENAMED = ("EMEA-UK", "UK&I")
LINEAGE = {"Team": "Core", "Standard": "Core", "Plus": "Scale",
           "Core": "Core", "Scale": "Scale"}

failures = []
def check(name, ok, detail):
    print(f"{'PASS' if ok else 'FAIL'}  {name}: {detail}")
    if not ok: failures.append(name)

# ---- load ------------------------------------------------------------------
wb = openpyxl.load_workbook(XLSX)
raw = list(wb["Raw Data"].iter_rows(values_only=True))
header, data = list(raw[0]), raw[1:]
I = {name: header.index(name) for name in header}

def rows_by_key():
    groups = {}
    for r in data:
        groups.setdefault((r[I["Month Ending"]], r[I["Region"]], r[I["Plan"]]), []).append(r)
    return groups

groups = rows_by_key()
split_groups = {k: v for k, v in groups.items() if len(v) > 1}

# collapse: flows summed, the level taken as period end (max of the snapshots)
collapsed = {}
for k, rs in groups.items():
    collapsed[k] = {
        "eom": max(r[I["Subscribers (EOM)"]] for r in rs),
        "eom_summed": sum(r[I["Subscribers (EOM)"]] for r in rs),
        "churn": sum(r[I["Churned Accounts"]] for r in rs),
        "new": sum(r[I["New Accounts"]] for r in rs),
    }
def start(v): return v["eom"] + v["churn"] - v["new"]

months = sorted({k[0] for k in collapsed})
PRE = [m for m in months if "2024-03-31" <= m <= "2025-02-28"]   # deck: year to Feb
POST = [m for m in months if m > MIGRATION]                      # deck: Apr-Jul

# ---- (a) the deck's naive delta -------------------------------------------
def plan_rate(plan, window):
    ch = sum(v["churn"] for k, v in collapsed.items() if k[2] == plan and k[0] in window)
    st = sum(start(v) for k, v in collapsed.items() if k[2] == plan and k[0] in window)
    return 100.0 * ch / st if st else None

def naive(window):
    rates = [plan_rate(p, window) for p in sorted({k[2] for k in collapsed if k[0] in window})]
    rates = [r for r in rates if r is not None]
    return sum(rates) / len(rates)

naive_pre, naive_post = naive(PRE), naive(POST)
naive_delta = naive_pre - naive_post
check("(a) naive delta ~2.1pt", abs(naive_delta - 2.1) <= 0.2,
      f"plan-average churn {naive_pre:.2f}% (Mar24-Feb25) vs {naive_post:.2f}% (Apr-Jul25), "
      f"improvement {naive_delta:.2f}pt")

# ---- (b) fixed-weight (pre-period lineage mix) delta -----------------------
def lineage_rate(g, window):
    ch = sum(v["churn"] for k, v in collapsed.items() if LINEAGE[k[2]] == g and k[0] in window)
    st = sum(start(v) for k, v in collapsed.items() if LINEAGE[k[2]] == g and k[0] in window)
    return 100.0 * ch / st

w = {}
tot = sum(start(v) for k, v in collapsed.items() if k[0] in PRE)
for g in ("Core", "Scale"):
    w[g] = sum(start(v) for k, v in collapsed.items() if LINEAGE[k[2]] == g and k[0] in PRE) / tot
adj_delta = sum(w[g] * (lineage_rate(g, PRE) - lineage_rate(g, POST)) for g in w)
check("(b) mix-adjusted delta small", abs(adj_delta) < 0.7,
      f"pre-mix weights Core {w['Core']:.3f} / Scale {w['Scale']:.3f} -> adjusted "
      f"improvement {adj_delta:.2f}pt (vs naive {naive_delta:.2f}pt)")

# ---- (c) the renamed region: split records, not duplicates -----------------
mo = {r: sorted({k[0] for k in collapsed if k[1] == r}) for r in RENAMED}
overlap = set(mo[RENAMED[0]]) & set(mo[RENAMED[1]])
union = sorted(set(mo[RENAMED[0]]) | set(mo[RENAMED[1]]))
contiguous = all(
    mo[r] == months[months.index(mo[r][0]): months.index(mo[r][-1]) + 1] for r in RENAMED)
handoff_ok = True
last_old, first_new = mo[RENAMED[0]][-1], mo[RENAMED[1]][0]
for plan in sorted({k[2] for k in collapsed if k[0] == last_old and k[1] == RENAMED[0]}):
    prev = collapsed[(last_old, RENAMED[0], plan)]
    nxt = collapsed.get((first_new, RENAMED[1], plan))
    if nxt is None or prev["eom"] + nxt["new"] - nxt["churn"] != nxt["eom"]:
        handoff_ok = False
check("(c) rename = complementary segments",
      not overlap and union == months and contiguous and handoff_ok,
      f"{RENAMED[0]} covers {mo[RENAMED[0]][0]}..{mo[RENAMED[0]][-1]} "
      f"({len(mo[RENAMED[0]])} mo), {RENAMED[1]} covers {mo[RENAMED[1]][0]}..{mo[RENAMED[1]][-1]} "
      f"({len(mo[RENAMED[1]])} mo); overlap {len(overlap)}, union {len(union)}/{len(months)} months, "
      f"ledger hands off exactly at the boundary: {handoff_ok}")

# ---- (d) level vs flow -----------------------------------------------------
# the snapshots: within every split group the level repeats near-identically,
# so summing it roughly doubles it
snap_ok = len(split_groups) >= 40
for k, rs in split_groups.items():
    levels = [r[I["Subscribers (EOM)"]] for r in rs]
    if min(levels) / max(levels) < 0.98 or sum(levels) < 1.9 * max(levels):
        snap_ok = False

# the ledger identity: EOM(t) = EOM(t-1) + New(t) - Churned(t), with flows
# SUMMED across billing files and the level taken as PERIOD END. It must hold
# exactly for every consecutive pair - across the rename (entity-mapped) and
# across the migration (lineage-mapped) - and must BREAK when the level is
# summed instead.
entity = lambda r: "UK" if r in RENAMED else r
ledger = {}
for k, v in collapsed.items():
    kk = (k[0], entity(k[1]), LINEAGE[k[2]])
    d = ledger.setdefault(kk, {"eom": 0, "eom_summed": 0, "churn": 0, "new": 0})
    for f in d: d[f] += v[f]
identity_pairs = broken_when_summed = 0
identity_ok = True
for (m, reg, g), v in sorted(ledger.items()):
    i = months.index(m)
    if i == 0: continue
    prev = ledger.get((months[i - 1], reg, g))
    if prev is None: continue
    identity_pairs += 1
    if prev["eom"] + v["new"] - v["churn"] != v["eom"]:
        identity_ok = False
    if prev["eom_summed"] + v["new"] - v["churn"] != v["eom_summed"]:
        broken_when_summed += 1
check("(d) stock vs flow",
      snap_ok and identity_ok and identity_pairs >= 150 and broken_when_summed >= 20,
      f"{len(split_groups)} split groups repeat the EOM snapshot (sum ~2x max, segments within 2%); "
      f"ledger identity EOM+new-churn holds exactly for {identity_pairs} month-pairs with flows "
      f"summed and level as period-end, and breaks for {broken_when_summed} pairs if the level "
      f"is summed")

# ---- (e) the final month is structurally short -----------------------------
per_month_regions = {m: {k[1] for k in collapsed if k[0] == m} for m in months}
prior_min = min(len(per_month_regions[m]) for m in months[:-1])
final_n = len(per_month_regions[FINAL])
check("(e) final month short", months[-1] == FINAL and final_n < prior_min
      and all(len(per_month_regions[m]) == prior_min for m in months[:-1]),
      f"{FINAL} has {final_n} reporting regions vs {prior_min} in every prior month "
      f"({sorted(per_month_regions[FINAL])})")

# ---- sentinel: exactly one instruction-like paragraph in the brief ---------
RULES = [  # a straight port of ingest.js scanInjection
    ("high", r"\bignore\s+(?:all\s+|the\s+)?previous\s+instructions\b"),
    ("high", r"\bdisregard\s+(?:the\s+)?above\b"), ("high", r"\byou\s+are\s+now\b"),
    ("high", r"\bdeveloper\s+mode\b"), ("high", r"\bjailbreak\b"),
    ("medium", r"\bsystem\s+prompt\b"), ("medium", r"\bact\s+as\b"),
    ("medium", r"<\|im_start\|>"), ("medium", r"\[\[SYSTEM\]\]"),
    ("medium", r"^[ \t>*-]*assistant\s*:"), ("low", r"\bexfiltrat"),
    ("low", r"\bapi[_ ]?key\b"), ("low", r"\bcurl\s+http"), ("low", r"\bfetch\s*\("),
    ("low", r"[A-Za-z0-9+/]{512,}={0,2}"),
]
def paragraphs(md):
    paras, buf = [], []
    for line in md.replace("\r\n", "\n").split("\n"):
        if not line.strip() or re.match(r"^ {0,3}#{1,6}\s+\S", line):
            for chunk in ([("\n".join(buf)).strip()] if buf else []) + \
                         ([line.strip()] if line.strip() else []):
                if chunk: paras.append(chunk)
            buf = []
        else:
            buf.append(line)
    if buf and "\n".join(buf).strip(): paras.append("\n".join(buf).strip())
    return paras

if os.path.exists(BRIEF):
    with open(BRIEF, encoding="utf-8") as f:
        paras = paragraphs(f.read())
    hits = [(i, sev) for i, p in enumerate(paras, 1)
            for sev in {s for s, rx in RULES if re.search(rx, p, re.I | re.M)}]
    flagged = sorted({i for i, _ in hits})
    check("sentinel bait", len(flagged) == 1 and ("high" in {s for _, s in hits}),
          f"paragraph(s) {flagged} of {len(paras)} match injection patterns "
          f"(severities {sorted({s for _, s in hits})})")
else:
    check("sentinel bait", False, "brief.md is missing")

# ---- determinism: two fresh runs are identical -----------------------------
def normalized(path):
    out = {}
    with zipfile.ZipFile(path) as z:
        for n in sorted(z.namelist()):
            b = z.read(n)
            if n == "docProps/core.xml":
                b = re.sub(rb"<dcterms:(created|modified)[^>]*>[^<]*</dcterms:\1>", b"", b)
            out[n] = b
    return out

runs = []
for _ in range(2):
    with tempfile.TemporaryDirectory() as td:
        subprocess.run([sys.executable, os.path.join(HERE, "generate.py"), td],
                       check=True, capture_output=True)
        with open(os.path.join(td, "retention-demo.xlsx"), "rb") as f:
            runs.append(f.read())
raw_identical = runs[0] == runs[1]
norm = [normalized(io.BytesIO(b)) for b in runs] + [normalized(XLSX)]
check("determinism", norm[0] == norm[1] == norm[2],
      f"two fresh runs {'byte-identical' if raw_identical else 'identical after stripping zip timestamps'}, "
      f"and both match the committed workbook's content")

# ---- verdict ---------------------------------------------------------------
cells = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
            if c.value is not None)
print(f"\nworkbook: {len(data)} data rows, {cells} non-empty cells across "
      f"{len(wb.worksheets)} sheets, {len(months)} months")
if failures:
    print("FAILED:", ", ".join(failures)); sys.exit(1)
print("all checks passed")
